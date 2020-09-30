import openpyxl

wb = openpyxl.load_workbook('C:\\Users\\sooa\\Desktop\\machine_learning\\Real estate valuation data set.xlsx')
sheet=wb['sheet1']
print(sheet.cell(row=2, column=1).value)
print("%s" % sheet.max_row)
